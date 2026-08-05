import csv
import json
import re
import tempfile
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook
from psycopg2.extras import execute_values

from database import conectar, criar_cursor


class ImportacaoProdutosService:
    MAX_LINHAS = 50000
    MAX_ARQUIVO = 20 * 1024 * 1024
    COLUNAS = {
        "nome": {"nome", "produto", "descricao", "descrição", "item"},
        "codigo_barras": {"codigo de barras", "código de barras", "codigo", "código", "ean", "gtin", "sku"},
        "preco": {"preco", "preço", "preco venda", "preço venda", "valor", "valor venda"},
        "estoque": {"estoque", "saldo", "quantidade", "qtd", "estoque atual"},
    }

    @staticmethod
    def _texto(valor):
        return str(valor or "").strip()

    @classmethod
    def _cabecalho(cls, valor):
        texto = cls._texto(valor).lower()
        texto = re.sub(r"[_\-]+", " ", texto)
        return re.sub(r"\s+", " ", texto).strip()

    @staticmethod
    def _decimal(valor):
        if valor is None or str(valor).strip() == "":
            return Decimal("0")
        if isinstance(valor, (int, float, Decimal)):
            return Decimal(str(valor))
        texto = re.sub(r"[^0-9,.-]", "", str(valor).strip())
        if "," in texto:
            texto = texto.replace(".", "").replace(",", ".")
        elif texto.count(".") > 1:
            texto = texto.replace(".", "")
        try:
            return Decimal(texto)
        except InvalidOperation as erro:
            raise ValueError("valor numérico inválido") from erro

    @classmethod
    def _linhas_arquivo(cls, nome, dados):
        extensao = Path(nome).suffix.lower()
        if extensao == ".xlsx":
            planilha = load_workbook(BytesIO(dados), read_only=True, data_only=True)
            aba = planilha.active
            return [list(linha) for linha in aba.iter_rows(values_only=True)]
        if extensao == ".csv":
            texto = dados.decode("utf-8-sig", errors="replace")
            amostra = texto[:4096]
            try:
                dialeto = csv.Sniffer().sniff(amostra, delimiters=";,\t")
            except csv.Error:
                dialeto = csv.excel
                dialeto.delimiter = ";"
            return list(csv.reader(StringIO(texto), dialeto))
        raise ValueError("Envie uma planilha XLSX ou CSV.")

    @classmethod
    def analisar(cls, nome, dados):
        if not dados:
            raise ValueError("O arquivo está vazio.")
        if len(dados) > cls.MAX_ARQUIVO:
            raise ValueError("O arquivo ultrapassa o limite de 20 MB.")
        linhas = cls._linhas_arquivo(nome, dados)
        if len(linhas) < 2:
            raise ValueError("A planilha precisa ter cabeçalho e ao menos um produto.")
        indice_cabecalho = None
        mapa = {}
        for indice_linha, candidata in enumerate(linhas[:10]):
            cabecalhos = [cls._cabecalho(valor) for valor in candidata]
            mapa_candidato = {}
            for campo, alternativas in cls.COLUNAS.items():
                for indice, coluna in enumerate(cabecalhos):
                    if coluna in alternativas:
                        mapa_candidato[campo] = indice
                        break
            if all(campo in mapa_candidato for campo in ("nome", "preco", "estoque")):
                indice_cabecalho = indice_linha
                mapa = mapa_candidato
                break

        if indice_cabecalho is None:
            raise ValueError(
                "Não encontrei o cabeçalho. Use as colunas Nome, Preço e Estoque."
            )

        linhas_produtos = linhas[indice_cabecalho + 1:]
        if len(linhas_produtos) > cls.MAX_LINHAS:
            raise ValueError("O limite é de 50.000 produtos por importação.")
        faltantes = [campo for campo in ("nome", "preco", "estoque") if campo not in mapa]
        if faltantes:
            raise ValueError("Colunas obrigatórias ausentes: " + ", ".join(faltantes) + ".")

        produtos, erros, codigos = [], [], set()
        for numero, linha in enumerate(
            linhas_produtos,
            start=indice_cabecalho + 2,
        ):
            if not any(cls._texto(valor) for valor in linha):
                continue
            try:
                def valor(campo):
                    indice = mapa.get(campo)
                    return linha[indice] if indice is not None and indice < len(linha) else ""

                nome_produto = cls._texto(valor("nome"))
                codigo = cls._texto(valor("codigo_barras")).lstrip("'")
                preco = cls._decimal(valor("preco"))
                estoque_decimal = cls._decimal(valor("estoque"))

                if not nome_produto:
                    raise ValueError("produto sem nome")
                if preco < 0 or estoque_decimal < 0:
                    raise ValueError("preço e estoque não podem ser negativos")
                if estoque_decimal != estoque_decimal.to_integral_value():
                    raise ValueError("estoque precisa ser um número inteiro")
                if codigo and codigo in codigos:
                    raise ValueError("código de barras repetido na planilha")
                if codigo:
                    codigos.add(codigo)
                produtos.append({
                    "linha": numero,
                    "nome": nome_produto[:180],
                    "codigo_barras": codigo[:80],
                    "preco": str(preco.quantize(Decimal("0.01"))),
                    "estoque": int(estoque_decimal),
                })
            except ValueError as erro:
                erros.append({"linha": numero, "erro": str(erro)})

        if not produtos:
            raise ValueError("Nenhum produto válido foi encontrado.")
        total_preenchidas = sum(
            1 for linha in linhas_produtos
            if any(cls._texto(valor) for valor in linha)
        )
        return {
            "produtos": produtos,
            "erros": erros,
            "total_linhas": total_preenchidas,
        }

    @staticmethod
    def salvar_temporario(token, resultado):
        caminho = Path(tempfile.gettempdir()) / f"nexus_importacao_{token}.json"
        caminho.write_text(json.dumps(resultado, ensure_ascii=False), encoding="utf-8")

    @staticmethod
    def carregar_temporario(token):
        caminho = Path(tempfile.gettempdir()) / f"nexus_importacao_{token}.json"
        if not caminho.exists():
            raise ValueError("A prévia expirou. Envie a planilha novamente.")
        return json.loads(caminho.read_text(encoding="utf-8"))

    @staticmethod
    def remover_temporario(token):
        caminho = Path(tempfile.gettempdir()) / f"nexus_importacao_{token}.json"
        caminho.unlink(missing_ok=True)

    @classmethod
    def importar(cls, empresa_id, produtos, estrategia="atualizar"):
        conn = conectar()
        cursor = criar_cursor(conn)
        try:
            codigos = [p["codigo_barras"] for p in produtos if p["codigo_barras"]]
            existentes = {}
            if codigos:
                cursor.execute(
                    "SELECT id, codigo_barras FROM produtos WHERE empresa_id = %s AND codigo_barras = ANY(%s)",
                    (empresa_id, codigos),
                )
                existentes = {str(item["codigo_barras"]): item["id"] for item in cursor.fetchall()}

            inserir, atualizar, ignorados = [], [], 0
            for produto in produtos:
                codigo = produto["codigo_barras"]
                registro = (
                    produto["nome"], Decimal(produto["preco"]), produto["estoque"],
                    codigo or None, "direto",
                )
                if codigo and codigo in existentes:
                    if estrategia == "ignorar":
                        ignorados += 1
                    else:
                        atualizar.append(registro + (existentes[codigo], empresa_id))
                else:
                    inserir.append(registro + (empresa_id,))

            if atualizar:
                execute_values(cursor, """
                    UPDATE produtos AS p SET nome = v.nome, preco = v.preco,
                    estoque = v.estoque, codigo_barras = v.codigo,
                    setor_preparo = v.setor
                    FROM (VALUES %s) AS v(nome, preco, estoque, codigo, setor, id, empresa_id)
                    WHERE p.id = v.id AND p.empresa_id = v.empresa_id
                """, atualizar, page_size=500)
            if inserir:
                execute_values(cursor, """
                    INSERT INTO produtos
                    (nome, preco, estoque, codigo_barras, setor_preparo, empresa_id)
                    VALUES %s
                """, inserir, page_size=500)
            conn.commit()
            return {"inseridos": len(inserir), "atualizados": len(atualizar), "ignorados": ignorados}
        except Exception:
            conn.rollback()
            raise
        finally:
            cursor.close()
            conn.close()
